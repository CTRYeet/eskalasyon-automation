import streamlit as st
import pandas as pd
import numpy as np
import requests
import xml.etree.ElementTree as ET
from datetime import datetime, date
import plotly.express as px
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------
# 0. SAYFA YAPILANDIRMASI & SESSION STATE
# ---------------------------------------------------------
st.set_page_config(
    page_title="LC WAIKIKI - Tedarikçi Eskalasyon Panel",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
    <style>
    .stAppDeployButton { display: none !important; }
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }
    </style>
""", unsafe_allow_html=True)

if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0

if "edited_data" not in st.session_state:
    st.session_state.edited_data = None


# ---------------------------------------------------------
# 1. CANLI TCMB KURU FONKSİYONLARI
# ---------------------------------------------------------
@st.cache_data(ttl=300)
def get_tcmb_live_usd():
    """TCMB'nin bugünkü resmi gösterge kurunu XML üzerinden canlı çeker."""
    try:
        url = "https://www.tcmb.gov.tr/kurlar/today.xml"
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            root = ET.fromstring(response.content)
            for currency in root.findall('Currency'):
                if currency.get('CurrencyCode') == 'USD':
                    rate = float(currency.find('ForexBuying').text)
                    return round(rate, 4)
    except Exception:
        pass
    return 34.80

canli_usd_kuru = get_tcmb_live_usd()


# ---------------------------------------------------------
# 2. GELİŞMİŞ VE GÜVENLİ EXCEL PARSER MOTORU
# ---------------------------------------------------------
def make_columns_unique(cols):
    seen = {}
    new_cols = []
    for col in cols:
        col_str = str(col).strip() if pd.notna(col) else "Unnamed"
        if col_str in seen:
            seen[col_str] += 1
            new_cols.append(f"{col_str}.{seen[col_str]}")
        else:
            seen[col_str] = 0
            new_cols.append(col_str)
    return new_cols

def clean_summary_rows(df, supplier_col="Tedarikçi"):
    if supplier_col in df.columns:
        filter_mask = ~df[supplier_col].astype(str).str.lower().str.strip().str.contains("toplam|total|genel|ortalama|summary", na=False)
        return df[filter_mask].reset_index(drop=True)
    return df

def clean_numeric_value(val):
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).replace("₺", "").replace("TL", "").replace("%", "").strip()
    if "," in val_str and "." in val_str:
        val_str = val_str.replace(".", "").replace(",", ".")
    elif "," in val_str:
        val_str = val_str.replace(",", ".")
    try:
        return float(val_str)
    except ValueError:
        return 0.0

def find_table_in_sheet(df_raw):
    """Excel sayfasında tablonun başladığı doğru satırı otomatik bulur."""
    for r in range(min(30, len(df_raw))):
        row_vals = [str(x).lower().strip() for x in df_raw.iloc[r, :].values if pd.notna(x)]
        row_str = " ".join(row_vals)
        if any(k in row_str for k in ["eskalasyon id", "tedarikçi", "firma", "hizmet tipi", "kategori", "sözleşme", "bütçe", "tutar", "mevcut birim fiyat"]):
            df = df_raw.iloc[r + 1:].reset_index(drop=True)
            df.columns = df_raw.iloc[r].values
            df.columns = [str(c).strip() if pd.notna(c) else f"Unnamed_{i}" for i, c in enumerate(df.columns)]
            return df
    return None

def universal_excel_parser(uploaded_file, target_sheet="Eskalasyon_Kayitlari"):
    """Excel sayfalarını akıllı eşleştirme ve ağırlık birleştirme ile okur."""
    xls = pd.ExcelFile(uploaded_file)
    sheet_names = xls.sheet_names
    
    main_df = None
    weights_df = None
    loaded_sheet_name = ""

    sheets_to_check = [target_sheet] if target_sheet in sheet_names else sheet_names
    
    for sheet in sheets_to_check:
        df_raw = pd.read_excel(uploaded_file, sheet_name=sheet, header=None)
        if df_raw.empty:
            continue
        
        parsed_df = find_table_in_sheet(df_raw)
        if parsed_df is not None:
            cols_lower = [str(c).lower() for c in parsed_df.columns]
            if any(k in " ".join(cols_lower) for k in ["eskalasyon id", "tedarikçi", "mevcut birim fiyat", "tutar", "bütçe"]) and main_df is None:
                main_df = parsed_df
                loaded_sheet_name = sheet
            elif any(k in " ".join(cols_lower) for k in ["hizmet tipi kodu", "tüfe ağırlığı", "motorin ağırlığı", "asgari ücret ağırlığı"]):
                weights_df = parsed_df

    if weights_df is None:
        for sheet in sheet_names:
            if sheet != loaded_sheet_name:
                df_raw = pd.read_excel(uploaded_file, sheet_name=sheet, header=None)
                parsed_df = find_table_in_sheet(df_raw)
                if parsed_df is not None:
                    cols_lower = [str(c).lower() for c in parsed_df.columns]
                    if any(k in " ".join(cols_lower) for k in ["hizmet tipi", "tüfe ağırlığı", "motorin ağırlığı", "ağırlık"]):
                        weights_df = parsed_df
                        break

    if main_df is None:
        main_df = pd.read_excel(uploaded_file, sheet_name=sheet_names[0])
        loaded_sheet_name = sheet_names[0]

    column_mapping = {}
    used_targets = set()

    for col in main_df.columns:
        c_clean = str(col).lower().replace("_", " ").replace("-", " ")
        target = col

        if any(k in c_clean for k in ["tedarikçi kuralı", "tedarikçi kodu", "supplier code"]) and "Tedarikçi_Kodu" not in used_targets:
            target = "Tedarikçi_Kodu"
        elif any(k in c_clean for k in ["tedarikçi adı", "tedarikçi ünvanı", "tedarikçi", "firma", "supplier", "vendor"]) and "Tedarikçi" not in used_targets:
            target = "Tedarikçi"
        elif any(k in c_clean for k in ["hizmet tipi", "hizmet türü", "kategori", "hizmet", "category"]) and "Kategori" not in used_targets:
            target = "Kategori"
        elif any(k in c_clean for k in ["mevcut birim fiyat", "sözleşme tutarı", "tutar", "bütçe", "fiyat", "amount"]) and "Tutar_TL" not in used_targets:
            target = "Tutar_TL"
        elif any(k in c_clean for k in ["talep edilen birim fiyat", "talep edilen tutar", "talep fiyat"]) and "Talep_TL" not in used_targets:
            target = "Talep_TL"
        elif any(k in c_clean for k in ["asgari", "işçilik", "iscilik"]) and any(k in c_clean for k in ["ağırlık", "w", "oran", "%"]) and "W_İşçilik" not in used_targets:
            target = "W_İşçilik"
        elif any(k in c_clean for k in ["yakıt", "yakit", "motorin"]) and any(k in c_clean for k in ["ağırlık", "w", "oran", "%"]) and "W_Yakıt" not in used_targets:
            target = "W_Yakıt"
        elif any(k in c_clean for k in ["tüfe", "üfe", "ufe"]) and any(k in c_clean for k in ["ağırlık", "w", "oran", "%"]) and "W_ÜFE" not in used_targets:
            target = "W_ÜFE"
        elif any(k in c_clean for k in ["döviz", "doviz", "usd", "eur"]) and any(k in c_clean for k in ["ağırlık", "w", "oran", "%"]) and "W_Döviz" not in used_targets:
            target = "W_Döviz"

        if target != col:
            used_targets.add(target)
            column_mapping[col] = target

    main_df = main_df.rename(columns=column_mapping)
    main_df.columns = make_columns_unique(main_df.columns)

    if weights_df is not None:
        w_map = {}
        w_used = set()
        for col in weights_df.columns:
            c_clean = str(col).lower()
            if ("hizmet tipi adı" in c_clean or "hizmet adı" in c_clean or "kategori adı" in c_clean or "hizmet tipi" in c_clean or "kategori" in c_clean) and "kodu" not in c_clean and "Kategori" not in w_used:
                w_map[col] = "Kategori"
                w_used.add("Kategori")
            elif "asgari" in c_clean or "işçilik" in c_clean:
                w_map[col] = "W_İşçilik"
            elif "motorin" in c_clean or "yakıt" in c_clean:
                w_map[col] = "W_Yakıt"
            elif "yi-üfe" in c_clean or "yi üfe" in c_clean:
                w_map[col] = "W_YI_UFE"
            elif "tüfe" in c_clean or "üfe" in c_clean:
                w_map[col] = "W_TUFE"
            elif "döviz" in c_clean:
                w_map[col] = "W_Döviz"

        weights_df = weights_df.rename(columns=w_map)
        weights_df.columns = make_columns_unique(weights_df.columns)

        w_tufe = pd.to_numeric(weights_df["W_TUFE"], errors='coerce').fillna(0) if "W_TUFE" in weights_df.columns else 0
        w_yi_ufe = pd.to_numeric(weights_df["W_YI_UFE"], errors='coerce').fillna(0) if "W_YI_UFE" in weights_df.columns else 0
        weights_df["W_ÜFE"] = w_tufe + w_yi_ufe

        merge_cols = [c for c in ["W_İşçilik", "W_Yakıt", "W_ÜFE", "W_Döviz"] if c in weights_df.columns]
        if merge_cols and "Kategori" in weights_df.columns and "Kategori" in main_df.columns:
            weights_subset = weights_df[["Kategori"] + merge_cols].drop_duplicates(subset=["Kategori"])
            main_df = main_df.merge(weights_subset, on="Kategori", how="left", suffixes=('', '_merged'))
            for col in merge_cols:
                if f"{col}_merged" in main_df.columns:
                    main_df[col] = main_df[col].fillna(main_df[f"{col}_merged"])
                    main_df = main_df.drop(columns=[f"{col}_merged"])

    main_df.columns = make_columns_unique(main_df.columns)

    if "Tedarikçi" in main_df.columns:
        main_df = main_df[main_df["Tedarikçi"].notna() & (main_df["Tedarikçi"].astype(str).str.strip() != "")].reset_index(drop=True)
        main_df = clean_summary_rows(main_df, "Tedarikçi")

        if "Tutar_TL" in main_df.columns:
            main_df["Tutar_TL"] = main_df["Tutar_TL"].apply(clean_numeric_value)
        if "Talep_TL" in main_df.columns:
            main_df["Talep_TL"] = main_df["Talep_TL"].apply(clean_numeric_value)
            
        for w_col in ["W_İşçilik", "W_Yakıt", "W_ÜFE", "W_Döviz"]:
            if w_col in main_df.columns:
                main_df[w_col] = main_df[w_col].apply(clean_numeric_value)

    for date_col in ['Sözleşme Başlangıç', 'Eskalasyon Tarihi']:
        if date_col in main_df.columns:
            main_df[date_col] = pd.to_datetime(main_df[date_col], errors='coerce').dt.strftime('%Y-%m-%d').fillna('')

    return main_df, loaded_sheet_name


# ---------------------------------------------------------
# 3. SIDEBAR & STRATEJİK YENİ MODÜLLER (BÜTÇE & HEDEF ORAN)
# ---------------------------------------------------------
with st.sidebar:
    st.metric(
        label="💱 USD/TRY (TCMB Canlı Kur)",
        value=f"{canli_usd_kuru:.2f} TL"
    )
    st.caption("🏛️ **Kaynak:** Türkiye Cumhuriyeti Merkez Bankası (TCMB)")
    st.divider()

    st.markdown("### ⚙️ Hesaplama Modu")
    calc_mode = st.radio(
        "Mod Seçimi",
        ["Excel Verisi (Varsayılan)", "Simülasyon / Manuel Mod"],
        help="Excel Verisi modu satır bazlı orijinal endeksleri okur. Simülasyon modu manuel değerleri uygular."
    )

    st.divider()
    st.markdown("### 📊 Manuel / Simülasyon Endeksleri")
    
    t0_iscilik = st.number_input("👷 T0 İşçilik (TL)", value=20002.50, step=100.0)
    t1_iscilik = st.number_input("👷 T1 İşçilik (TL)", value=22104.00, step=100.0)
    ratio_iscilik = t1_iscilik / t0_iscilik if t0_iscilik > 0 else 1.0

    t0_yakit = st.number_input("⛽ T0 Yakıt (TL/Lt)", value=42.50, step=0.5)
    t1_yakit = st.number_input("⛽ T1 Yakıt (TL/Lt)", value=48.90, step=0.5)
    ratio_yakit = t1_yakit / t0_yakit if t0_yakit > 0 else 1.0

    t0_ufe = st.number_input("🏭 T0 ÜFE", value=3250.00, step=10.0)
    t1_ufe = st.number_input("🏭 T1 ÜFE", value=3580.00, step=10.0)
    ratio_ufe = t1_ufe / t0_ufe if t0_ufe > 0 else 1.0

    t0_doviz = st.number_input("💵 T0 USD/TRY", value=32.50, step=0.1)
    t1_doviz = st.number_input("💵 T1 USD/TRY", value=34.80, step=0.1)
    ratio_doviz = t1_doviz / t0_doviz if t0_doviz > 0 else 1.0

    if calc_mode == "Simülasyon / Manuel Mod":
        pct_iscilik = (ratio_iscilik - 1) * 100
        pct_yakit = (ratio_yakit - 1) * 100
        pct_ufe = (ratio_ufe - 1) * 100
        pct_doviz = (ratio_doviz - 1) * 100

        st.caption(f"👷 **İşçilik Oranı:** {ratio_iscilik:.4f} (+%{pct_iscilik:.2f})")
        st.caption(f"⛽ **Yakıt Oranı:** {ratio_yakit:.4f} (+%{pct_yakit:.2f})")
        st.caption(f"🏭 **ÜFE Oranı:** {ratio_ufe:.4f} (+%{pct_ufe:.2f})")
        st.caption(f"💵 **Döviz Oranı:** {ratio_doviz:.4f} (+%{pct_doviz:.2f})")

    st.divider()

    st.markdown("### 🎯 Pazarlık & Bütçe Yönetimi")
    
    # STRATEJİK MODÜL 1: Dynamic Bargaining Target Slider
    hedef_anlasma_orani = st.slider(
        "Hedef Anlaşma Oranı (%)", 
        min_value=0.0, 
        max_value=50.0, 
        value=30.0, 
        step=0.5,
        help="Pazarlık masasında hedeflenen nihai eskalasyon artış oranı."
    )

    # STRATEJİK MODÜL 3: Budget Limit Input
    donemlik_butce = st.number_input(
        "Dönemlik Eskalasyon Bütçesi (TL)", 
        value=850000.0, 
        step=25000.0,
        format="%.2f"
    )

    st.divider()
    tolerans = st.slider("Tolerans Limiti (%)", min_value=0.0, max_value=50.0, value=29.0, step=0.1)
    rapor_donemi = st.text_input("Raporlama Dönemi", value="Q3 - 2026")
    safe_file_name = "".join(c for c in rapor_donemi if c.isalnum() or c in (" ", "_", "-")).strip().replace(" ", "_")
    st.divider()

    initial_data = {
        "Tedarikçi_Kodu": ["SUP-001", "SUP-002", "SUP-003", "SUP-004", "SUP-005", "SUP-006", "SUP-007", "SUP-008", "SUP-009", "SUP-010"],
        "Tedarikçi": [
            "Horoz Lojistik A.Ş.", "Netlog Lojistik Hizmetleri", "Ekol Lojistik A.Ş.", "Mars Logistics Group",
            "Sertrans Logistics", "Ceva Lojistik Limited", "Reysaş Taşımacılık", "Barsan Global Lojistik",
            "Fevzi Gandur Logistics", "OMS Lojistik A.Ş."
        ],
        "Kategori": [
            "FTL Komple Taşıma", "LTL Parsiyel", "Cross-Docking", "Depolama",
            "FTL Komple Taşıma", "Gümrükleme", "LTL Parsiyel", "FTL Komple Taşıma",
            "Depolama", "Cross-Docking"
        ],
        "Tutar_TL": [146238.52, 204850.03, 113329.33, 188723.42, 318232.42, 510084.21, 386379.74, 254010.60, 418441.90, 458048.67],
        "Talep_TL": [190000.00, 270000.00, 150000.00, 250000.00, 415000.00, 675000.00, 505000.00, 335000.00, 545000.00, 600000.00],
        "W_İşçilik": [0.35, 0.30, 0.30, 0.25, 0.25, 0.20, 0.25, 0.20, 0.20, 0.25],
        "W_Yakıt":   [0.35, 0.35, 0.30, 0.30, 0.30, 0.30, 0.25, 0.30, 0.25, 0.30],
        "W_ÜFE":     [0.20, 0.25, 0.30, 0.35, 0.35, 0.40, 0.40, 0.40, 0.45, 0.35],
        "W_Döviz":   [0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10]
    }
    df_sample = pd.DataFrame(initial_data)

    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        df_sample.to_excel(writer, index=False, sheet_name='Eskalasyon_Sablonu')
    excel_bytes = output_buffer.getvalue()

    st.download_button(
        label="📥 Şablon Excel İndir",
        data=excel_bytes,
        file_name=f"Eskalasyon_Sablonu_{safe_file_name}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    if st.button("🗑️ Verileri Temizle", use_container_width=True):
        st.session_state.uploader_key += 1
        st.session_state.edited_data = pd.DataFrame(columns=df_sample.columns)
        st.rerun()


# ---------------------------------------------------------
# 4. ANA İÇERİK & VERİ YÜKLEME
# ---------------------------------------------------------
st.title("📦 LC WAIKIKI - Tedarikçi Eskalasyon & Fiyat Güncelleme")
st.caption("Ağırlıklı endeks eskalasyon modeli ve sözleşme analiz paneli")

active_df = pd.DataFrame()

tab1, tab2 = st.tabs(["📝 Canlı Excel Sheet (Maliyet Ağırlıkları)", "📁 Harici Excel Yükle"])

with tab1:
    st.markdown("##### 💡 *Maliyet ağırlıklarının (W_İşçilik, W_Yakıt, W_ÜFE, W_Döviz) toplamı satır bazında 1.00 (%100) olmalıdır.*")
    
    if st.session_state.edited_data is None:
        st.session_state.edited_data = df_sample.copy()

    edited_df = st.data_editor(
        st.session_state.edited_data,
        num_rows="dynamic",
        use_container_width=True,
        height=320,
        key="data_editor_widget"
    )
    active_df = edited_df.copy()

with tab2:
    st.markdown("### 📁 Excel Dosyası Yükleyin")
    uploaded_file = st.file_uploader(
        "Sözleşme verilerini içeren Excel dosyasını seçin", 
        type=["xlsx", "xls"],
        key=f"file_uploader_{st.session_state.uploader_key}"
    )
    
    if uploaded_file is not None:
        try:
            cleaned_excel, loaded_sheet = universal_excel_parser(uploaded_file, target_sheet="Eskalasyon_Kayitlari")
            st.success("Eskalasyon Kayıtları ve ilişkili veriler başarıyla ayrıştırıldı!")
            
            with st.expander("🔍 Gelişmiş Ham Veri Tablosunu Görüntüle", expanded=False):
                st.dataframe(cleaned_excel, use_container_width=True, height=250)
                
            active_df = cleaned_excel.copy()
        except Exception as e:
            st.error(f"Excel okunurken hata oluştu: {e}")

st.divider()


# ---------------------------------------------------------
# 5. STRATEJİK MODÜL 2: HİZMET KATEGORİSİ FİLTRESİ (MULTISELECT)
# ---------------------------------------------------------
if not active_df.empty:
    if "Kategori" in active_df.columns:
        mevcut_kategoriler = sorted(list(active_df["Kategori"].dropna().unique()))
        
        selected_kategoriler = st.multiselect(
            "🚚 Hizmet Kategorisi Filtrele:",
            options=mevcut_kategoriler,
            default=mevcut_kategoriler,
            help="Tedarikçileri ve analizi seçilen hizmet kategorilerine göre filtrelenmiş olarak günceller."
        )

        if selected_kategoriler:
            active_df = active_df[active_df["Kategori"].isin(selected_kategoriler)].reset_index(drop=True)
        else:
            st.warning("⚠️ En az bir hizmet kategorisi seçilmelidir.")
            active_df = active_df.iloc[0:0]


# ---------------------------------------------------------
# 6. HESAPLAMA MOTORU & DİNAMİK METRİKLER
# ---------------------------------------------------------
weight_cols = ["W_İşçilik", "W_Yakıt", "W_ÜFE", "W_Döviz"]

if not active_df.empty:
    tedarikci_col = "Tedarikçi" if "Tedarikçi" in active_df.columns else active_df.columns[0]
    active_df = clean_summary_rows(active_df, tedarikci_col)

    tutar_col = "Tutar_TL" if "Tutar_TL" in active_df.columns else active_df.columns[0]
    active_df[tutar_col] = pd.to_numeric(active_df[tutar_col], errors="coerce").fillna(0)

    for col in weight_cols:
        if col in active_df.columns:
            active_df[col] = pd.to_numeric(active_df[col], errors="coerce").fillna(0)
        else:
            active_df[col] = 0.0

    def process_weights(row):
        w_sum = sum(row[col] for col in weight_cols)
        if w_sum > 1.5:  
            norm_w = [row[col] / 100.0 for col in weight_cols]
        elif w_sum > 0:  
            norm_w = [row[col] / w_sum for col in weight_cols]
        else:
            norm_w = [0.25, 0.25, 0.25, 0.25]

        return pd.Series(norm_w, index=weight_cols)

    weight_processed = active_df.apply(process_weights, axis=1)

    def calculate_row_k(row, idx):
        if calc_mode == "Excel Verisi (Varsayılan)" and "Referans TÜFE (t0)" in row and "Uygulama TÜFE (t1)" in row and pd.notna(row["Referans TÜFE (t0)"]) and row["Referans TÜFE (t0)"] > 0:
            r_tufe = float(row["Uygulama TÜFE (t1)"]) / float(row["Referans TÜFE (t0)"])
            r_yi_ufe = (float(row["Uygulama Yİ-ÜFE (t1)"]) / float(row["Referans Yİ-ÜFE (t0)"])) if "Referans Yİ-ÜFE (t0)" in row and pd.notna(row["Referans Yİ-ÜFE (t0)"]) and row["Referans Yİ-ÜFE (t0)"] > 0 else r_tufe
            r_yakit = (float(row["Uygulama Motorin (t1)"]) / float(row["Referans Motorin (t0)"])) if "Referans Motorin (t0)" in row and pd.notna(row["Referans Motorin (t0)"]) and row["Referans Motorin (t0)"] > 0 else ratio_yakit
            r_iscilik = (r_tufe + r_yi_ufe) / 2.0
            r_doviz = ratio_doviz
            
            k_val = (weight_processed.loc[idx, "W_İşçilik"] * r_iscilik) + \
                    (weight_processed.loc[idx, "W_Yakıt"] * r_yakit) + \
                    (weight_processed.loc[idx, "W_ÜFE"] * r_tufe) + \
                    (weight_processed.loc[idx, "W_Döviz"] * r_doviz)
            return k_val
        else:
            k_val = (weight_processed.loc[idx, "W_İşçilik"] * ratio_iscilik) + \
                    (weight_processed.loc[idx, "W_Yakıt"] * ratio_yakit) + \
                    (weight_processed.loc[idx, "W_ÜFE"] * ratio_ufe) + \
                    (weight_processed.loc[idx, "W_Döviz"] * ratio_doviz)
            return k_val

    active_df["Katsayı (K)"] = [calculate_row_k(row, idx) for idx, row in active_df.iterrows()]
    active_df["Formül_Tutar_TL"] = active_df[tutar_col] * active_df["Katsayı (K)"]

    if "Talep_TL" not in active_df.columns:
        active_df["Talep_TL"] = active_df["Formül_Tutar_TL"]

    agg_dict = {
        tutar_col: "sum",
        "Talep_TL": "sum",
        "Formül_Tutar_TL": "sum"
    }

    if "Tedarikçi_Kodu" in active_df.columns:
        agg_dict["Tedarikçi_Kodu"] = "first"

    summary_df = active_df.groupby(tedarikci_col, as_index=False).agg(agg_dict)

    summary_df["Talep_Artış_%"] = ((summary_df["Talep_TL"] - summary_df[tutar_col]) / summary_df[tutar_col]) * 100.0
    summary_df["Formül_Artış_%"] = ((summary_df["Formül_Tutar_TL"] - summary_df[tutar_col]) / summary_df[tutar_col]) * 100.0
    summary_df["Müzakere_Farkı_TL"] = summary_df["Talep_TL"] - summary_df["Formül_Tutar_TL"]

    # Tolerans Sınırı Durumu (%29)
    tol_col_key = f"Tolerans Sınır Durumu (%{tolerans:.0f})"
    summary_df[tol_col_key] = summary_df["Formül_Artış_%"].apply(
        lambda x: "Tolerans İçi (Uygun)" if x <= tolerans else "Tolerans Aşımı (İncelemeli)"
    )

    # Hiyerarşik Önerilen Aksiyon Mantığı
    def determine_bargain_action(row):
        formul_pct = row["Formül_Artış_%"]
        talep_pct = row["Talep_Artış_%"]
        diff_pct = talep_pct - formul_pct

        if formul_pct <= tolerans and diff_pct <= 5.0:
            return "Doğrudan Onayla (Kabul Edilebilir)"
        elif formul_pct <= tolerans and diff_pct > 5.0:
            return "Talep Yüksek (Formülden Onayla)"
        else: # formul_pct > tolerans
            return "Pazarlığa Git (Revize İstenecek)"

    summary_df["Önerilen Pazarlık Aksiyonu"] = summary_df.apply(determine_bargain_action, axis=1)

    # Hacim Ağırlıklı Portföy Metrikleri
    mevcut_butce = summary_df[tutar_col].sum()
    guncel_butce = summary_df["Formül_Tutar_TL"].sum()
    toplam_talep_tl = summary_df["Talep_TL"].sum()
    fark = guncel_butce - mevcut_butce
    fark_yuzde = ((guncel_butce - mevcut_butce) / mevcut_butce * 100.0) if mevcut_butce > 0 else 0.0

    # STRATEJİK MODÜL 1 FORMÜLÜ: Potansiyel Şirket Tasarrufu
    hedef_anlasma_butcesi = mevcut_butce * (1.0 + (hedef_anlasma_orani / 100.0))
    potansiyel_tasarruf = toplam_talep_tl - hedef_anlasma_butcesi

    # STRATEJİK MODÜL 3 FORMÜLÜ: Bütçe Aşım Riski
    formul_maliyet_farki = guncel_butce - mevcut_butce
    butce_asim_farki = formul_maliyet_farki - donemlik_butce

else:
    summary_df = pd.DataFrame()
    mevcut_butce = 0.0
    guncel_butce = 0.0
    toplam_talep_tl = 0.0
    fark = 0.0
    fark_yuzde = 0.0
    potansiyel_tasarruf = 0.0
    formul_maliyet_farki = 0.0
    butce_asim_farki = -donemlik_butce


# ---------------------------------------------------------
# 7. KPI KARTLARI (YENİ STRATEJİK MODÜLLER İLE 5 KARTLI YAPI)
# ---------------------------------------------------------
kpi_col1, kpi_col2, kpi_col3, kpi_col4, kpi_col5 = st.columns(5)

with kpi_col1:
    st.markdown("##### Sözleşme Hacmi ($P_0$)")
    st.markdown(f"### ₺{mevcut_butce:,.2f}")

with kpi_col2:
    st.markdown("##### Formül Hacmi ($P_{\\text{calc}}$)")
    st.markdown(f"### ₺{guncel_butce:,.2f}")
    st.caption(f"🟢 **Formül Farkı:** +₺{fark:,.2f}")

with kpi_col3:
    st.markdown("##### Hacim Ağırlıklı Artış")
    st.markdown(f"### %{fark_yuzde:.2f}")
    if fark_yuzde <= tolerans:
        st.caption(f"🟢 Tolerans Sınırı (%{tolerans:.1f}) İçi")
    else:
        st.caption(f"🔴 Tolerans Aşımı! Limit: %{tolerans:.1f}")

with kpi_col4:
    # STRATEJİK MODÜL 1 KPI KARTI: Potansiyel Şirket Tasarrufu
    st.markdown("##### Potansiyel Tasarruf")
    st.markdown(f"### ₺{potansiyel_tasarruf:,.2f}")
    if potansiyel_tasarruf > 0:
        st.caption(f"🟢 Tedarikçi Talebine Göre **₺{potansiyel_tasarruf:,.2f}** Tasarruf İmkanı")
    else:
        st.caption(f"⚪ Hedef Oran Talepten Yüksek")

with kpi_col5:
    # STRATEJİK MODÜL 3 KPI KARTI: Eskalasyon Bütçe Riski
    st.markdown("##### Bütçe Riski Analizi")
    if butce_asim_farki > 0:
        st.markdown(f"### ₺{butce_asim_farki:,.2f}")
        st.caption(f"🔴 **₺{butce_asim_farki:,.2f} Bütçe Aşım Riski!**")
    else:
        kalan_marj = abs(butce_asim_farki)
        st.markdown(f"### ₺0.00")
        st.caption(f"🟢 **Bütçe İçi** (Kalan Marj: ₺{kalan_marj:,.2f})")

st.divider()


# ---------------------------------------------------------
# 8. İKİLİ ÇUBUK GRAFİK (SIDE-BY-SIDE BAR CHART)
# ---------------------------------------------------------
st.subheader("📊 Tedarikçi Bazlı Artış Karşılaştırması ve Tolerans Analizi")

if not summary_df.empty:
    chart_prep = summary_df[[tedarikci_col, "Talep_Artış_%", "Formül_Artış_%"]].copy()
    chart_prep[tedarikci_col] = chart_prep[tedarikci_col].astype(str)
    
    chart_prep = chart_prep.rename(columns={
        "Talep_Artış_%": "Tedarikçi Talep Artış %",
        "Formül_Artış_%": "Formül Hesaplanan Artış %"
    })

    chart_melted = chart_prep.melt(
        id_vars=[tedarikci_col],
        value_vars=["Tedarikçi Talep Artış %", "Formül Hesaplanan Artış %"],
        var_name="Metrik",
        value_name="Artış Oranı (%)"
    )

    fig = px.bar(
        chart_melted,
        x=tedarikci_col,
        y="Artış Oranı (%)",
        color="Metrik",
        barmode="group",
        color_discrete_map={
            "Tedarikçi Talep Artış %": "#E65100",        # Kırmızı / Turuncu
            "Formül Hesaplanan Artış %": "#1B365D"       # LC Waikiki Lacivert
        },
        text=chart_melted["Artış Oranı (%)"].apply(lambda x: f"%{x:.2f}"),
        labels={"Artış Oranı (%)": "Artış Oranı (%)", tedarikci_col: "Tedarikçi Firma"},
        title=f"Tedarikçi Talebi vs Formül Hesaplanan Artış Oranı (Tolerans Sınırı: %{tolerans:.1f})"
    )

    fig.add_hline(
        y=tolerans, 
        line_dash="dash", 
        line_color="#D32F2F", 
        line_width=2,
        annotation_text=f"Tolerans Sınırı (%{tolerans:.1f})", 
        annotation_position="top right"
    )

    fig.update_traces(
        textposition="outside"
    )

    fig.update_layout(
        template="plotly_white",
        height=480,
        legend_title_text="Karşılaştırma",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        xaxis=dict(
            title=dict(text="Tedarikçi Firma", font=dict(size=14)),
            tickfont=dict(size=12)
        ),
        yaxis=dict(
            title=dict(text="Artış Oranı (%)", font=dict(size=14)),
            tickfont=dict(size=12)
        )
    )

    st.plotly_chart(fig, use_container_width=True)


# ---------------------------------------------------------
# 9. GELİŞMİŞ EXCEL EXPORT (OPENPYXL İLE KURUMSAL STİL)
# ---------------------------------------------------------
st.divider()

with st.expander("📋 Tedarikçi Bazlı Eskalasyon Özet Tablosu", expanded=True):
    if not summary_df.empty:
        tol_col_name = f"Tolerans Sınır Durumu (%{tolerans:.0f})"
        
        cols_order = []
        if "Tedarikçi_Kodu" in summary_df.columns: cols_order.append("Tedarikçi_Kodu")
        cols_order.extend([
            tedarikci_col, tutar_col, "Talep_TL", "Talep_Artış_%", 
            "Formül_Tutar_TL", "Formül_Artış_%", "Müzakere_Farkı_TL", 
            tol_col_key, "Önerilen Pazarlık Aksiyonu"
        ])
        
        res_df = summary_df[[c for c in cols_order if c in summary_df.columns]].copy()
        
        res_df = res_df.rename(columns={
            "Tedarikçi_Kodu": "Tedarikçi Kodu",
            tedarikci_col: "Tedarikçi Unvanı",
            tutar_col: "Mevcut Tutar (P0)",
            "Talep_TL": "Tedarikçi Talebi (P1_Talep)",
            "Talep_Artış_%": "Talep Artış %",
            "Formül_Tutar_TL": "Formül Tutarı (Pcalc)",
            "Formül_Artış_%": "Formül Artış %",
            "Müzakere_Farkı_TL": "Müzakere Farkı (TL)",
            tol_col_key: tol_col_name
        })

        text_cols = [c for c in ["Tedarikçi Kodu", "Tedarikçi Unvanı", tol_col_name, "Önerilen Pazarlık Aksiyonu"] if c in res_df.columns]
        num_cols = [c for c in res_df.columns if c not in text_cols]

        formatted_styler = res_df.style\
            .format({
                "Mevcut Tutar (P0)": "₺{:,.2f}",
                "Tedarikçi Talebi (P1_Talep)": "₺{:,.2f}",
                "Talep Artış %": "%{:.2f}",
                "Formül Tutarı (Pcalc)": "₺{:,.2f}",
                "Formül Artış %": "%{:.2f}",
                "Müzakere Farkı (TL)": "₺{:,.2f}"
            })\
            .set_properties(subset=text_cols, **{'text-align': 'left'})\
            .set_properties(subset=num_cols, **{'text-align': 'right'})

        st.dataframe(
            formatted_styler,
            use_container_width=True,
            hide_index=True
        )

        excel_out_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_out_buffer, engine='openpyxl') as writer:
            res_df.to_excel(writer, index=False, sheet_name='Eskalasyon_Karsilastirma_Raporu')

        excel_out_buffer.seek(0)
        wb = openpyxl.load_workbook(excel_out_buffer)
        ws = wb['Eskalasyon_Karsilastirma_Raporu']

        header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid") # LCW Lacivert
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Segoe UI", size=10)
        
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Soft Yeşil
        green_font = Font(name="Segoe UI", size=10, color="006100")
        
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Soft Kırmızı
        red_font = Font(name="Segoe UI", size=10, color="9C0006")

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 28

        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = regular_font
                cell.border = thin_border
                col_header = headers[cell.column - 1]

                if col_header in ["Mevcut Tutar (P0)", "Tedarikçi Talebi (P1_Talep)", "Formül Tutarı (Pcalc)", "Müzakere Farkı (TL)"]:
                    cell.number_format = '₺#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                
                elif col_header in ["Talep Artış %", "Formül Artış %"]:
                    cell.number_format = '0.00"%"'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                if "Tolerans Sınır" in str(col_header):
                    cell_val = str(cell.value)
                    if "Tolerans İçi" in cell_val:
                        cell.fill = green_fill
                        cell.font = green_font
                    elif "Tolerans Aşımı" in cell_val:
                        cell.fill = red_fill
                        cell.font = red_font

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        final_buffer = io.BytesIO()
        wb.save(final_buffer)
        excel_res_bytes = final_buffer.getvalue()

        st.download_button(
            label="📥 Sonuç Raporunu İndir (.xlsx)",
            data=excel_res_bytes,
            file_name=f"LCW_Eskalasyon_Karsilastirma_Raporu_{safe_file_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )